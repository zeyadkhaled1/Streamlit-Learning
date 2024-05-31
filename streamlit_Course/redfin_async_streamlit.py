from curl_cffi import requests
from selectolax.parser import HTMLParser
import pandas as pd
import json_repair
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import asyncio
import time
from pandas import DataFrame
import os
def timing_decorator(func):
    async def wrapper_function(*args,**kwargs):
        start = time.perf_counter()
        result=await func(*args,**kwargs)
        end = time.perf_counter()
        x=end-start
        print(x)
        return result
    return wrapper_function   

def progress_bar(total=None,reset=False):
    if not hasattr(progress_bar, "progress"):
        progress_bar.progress = 1
    if reset :
        progress_bar.progress=1
        return
    percent = 100*(progress_bar.progress / float(total))
    bar = '█' * int(percent) + '-' * (100-int(percent))
    print(f"\r|{bar}| {percent:.2f}%", end='\r')
    progress_bar.progress += 1

state_abbreviations = {
    "ALABAMA": "AL",
    "ALASKA": "AK",
    "AMERICAN SAMOA": "AS",
    "ARIZONA": "AZ",
    "ARKANSAS": "AR",
    "CALIFORNIA": "CA",
    "COLORADO": "CO",
    "CONNECTICUT": "CT",
    "DELAWARE": "DE",
    "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL",
    "GEORGIA": "GA",
    "GUAM": "GU",
    "HAWAII": "HI",
    "IDAHO": "ID",
    "ILLINOIS": "IL",
    "INDIANA": "IN",
    "IOWA": "IA",
    "KANSAS": "KS",
    "KENTUCKY": "KY",
    "LOUISIANA": "LA",
    "MAINE": "ME",
    "MARYLAND": "MD",
    "MASSACHUSETTS": "MA",
    "MICHIGAN": "MI",
    "MINNESOTA": "MN",
    "MISSISSIPPI": "MS",
    "MISSOURI": "MO",
    "MONTANA": "MT",
    "NEBRASKA": "NE",
    "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH",
    "NEW JERSEY": "NJ",
    "NEW MEXICO": "NM",
    "NEW YORK": "NY",
    "NORTH CAROLINA": "NC",
    "NORTH DAKOTA": "ND",
    "NORTHERN MARIANA IS": "MP",
    "OHIO": "OH",
    "OKLAHOMA": "OK",
    "OREGON": "OR",
    "PENNSYLVANIA": "PA",
    "PUERTO RICO": "PR",
    "RHODE ISLAND": "RI",
    "SOUTH CAROLINA": "SC",
    "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN",
    "TEXAS": "TX",
    "UTAH": "UT",
    "VERMONT": "VT",
    "VIRGINIA": "VA",
    "VIRGIN ISLANDS": "VI",
    "WASHINGTON": "WA",
    "WEST VIRGINIA": "WV",
    "WISCONSIN": "WI",
    "WYOMING": "WY",
}

def write_data_to_excel(results: list[tuple[str, str, str, str]],counties:list[str]):
    workbook=Workbook()
    entry_data_sheet = workbook.create_sheet(title=f"{datetime.today().date().strftime('%Y-%m-%d')} result", index=0)
    entry_data_sheet.append(["County", "Date Ran","Solds Homes (30 Days)", "Listed Homes (All)", "Ratio Homes", "Solds (30 Days)","Listed (All)","Ratio"])
    for result,county in zip(results,counties):
            entry_data_sheet.append([county,datetime.today().date().strftime("%Y-%m-%d"),result[0],result[1],f"{round(int(result[0])/int(result[1])*100,1)}%",result[2],result[3],f"{round(int(result[2])/int(result[3])*100,1)}%"])
    workbook.save("result.xlsx")

def write_data_to_excel_2(results: list[tuple[tuple[str,str],tuple[str, str, str, str]]],counties:list[str]):
    list_of_data_to_write=[]
    for (county,result) in results:
        percentage_of_houses=f"{round(int(result[0])/int(result[1])*100,1)}%" if int(result[1])!=0 else "0%"
        percentage_of_lands=f"{round(int(result[2])/int(result[3])*100,1)}%" if int(result[3])!=0 else "0%"
        list_of_data_to_write.append([','.join(county),datetime.today().date().strftime("%Y-%m-%d"),result[0],result[1],percentage_of_houses,result[2],result[3],percentage_of_lands])
    # Load the workbook
    wb = load_workbook("CountyDemandMaster.xlsx")
    # Select the active sheet
    ws = wb['May 2024']
    # Get the starting row index
    start_row = 2
    # Get the starting column index
    start_col = 1
    black_font = Font(color="000000")
    # Iterate over the data and write it to the sheet
    for row_idx, row_data in enumerate(list_of_data_to_write, start=start_row):
        for col_idx, cell_value in enumerate(row_data, start=start_col):
            cell=ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = black_font 
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value is not None:  # Only change the font if the cell is not empty
            cell.font = black_font
    # Save the workbook
    wb.save("CountyDemandMaster.xlsx")

def data_to_dataframe(results: list[tuple[tuple[str,str],tuple[str, str, str, str]]],)->DataFrame:
    list_of_data_to_write=[]
    for (county,result) in results:
        percentage_of_houses=f"{round(int(result[0])/int(result[1])*100,1)}%" if int(result[1])!=0 else "0%"
        percentage_of_lands=f"{round(int(result[2])/int(result[3])*100,1)}%" if int(result[3])!=0 else "0%"
        list_of_data_to_write.append([','.join(county),datetime.today().date().strftime("%Y-%m-%d"),result[0],result[1],percentage_of_houses,result[2],result[3],percentage_of_lands])
    columns=["County", "Date Ran","Solds Homes (30 Days)", "Listed Homes (All)", "Ratio Homes", "Solds (30 Days)","Listed (All)","Ratio"]
    df = pd.DataFrame(list_of_data_to_write, columns=columns)
    return df

def read_counties_from_file(uploaded_file) -> list[str]:
    df = pd.read_excel(uploaded_file, header=None, sheet_name="May 2024")
    return df.iloc[1:, 0].tolist()


def split_county_to_tuple(list_of_counties: list[str]) -> list[tuple[str, str]]:
    list_of_tuples = []
    for county in list_of_counties:
        county = county.split(",")
        if len(county) == 2:
            county_name, state = county
            state = state.strip()
            county_name=county_name.replace(".","",1) if county_name[0]=='.' else county_name
            county_name = county_name.strip()
            list_of_tuples.append((county_name, state))
        else:
            print(f"Error: {county}")
    return list_of_tuples

def make_search_request(query:tuple[str,str],query_type:int):
    base_url = "https://www.redfin.com/stingray/do/location-autocomplete"
    if query_type==1:
        location=f"{query[0]}"
    elif query_type==2:
        location=f"{query[0]} {query[1]}"
    else:
        location=f"{query[0].split(' ')[0]}"
    params = {
        "location": location,
        "start": 0,
        "count": 10,
        "v": 2,
        "market": "utah",
        "al": 2,
        "iss": True,
        "ooa": True,
        "mrs": False,
        "region_id": 2358,
        "region_type": 5,
        "lat": 45.548625,
        "lng": -123.11470500000001,
        "includeAddressInfo": False,
    }
    response = requests.get(base_url, params=params,impersonate='chrome110')
    decoded_object:dict = json_repair.loads((response.text)[4:-1])
    return decoded_object['payload']['sections'][0]['rows']

def compare_and_get_link(json_results:dict,query:tuple[str,str]):
    for result in json_results:
        if ((query[0].lower() == result['name'].lower()) or(' '.join(query[0].split()[:-1]).lower() == result['name'].lower()) ) and result['subName'].split(",")[0].strip() == state_abbreviations[query[1].upper()]:
            return result['url']
        if ((query[0].lower() == result['name'].lower()) or(' '.join(query[0].split()[:-1]).lower() == result['name'].lower())) and (' '.join(query[0].split()[:-1]).lower() == result['subName'].split(",")[0].strip().lower()) and (state_abbreviations[query[1].upper()] == result['subName'].split(",")[1].strip()): 
            return result['url']
    return None

async def get_search_links_associated_with_search_tuple(list_of_tuples: list[tuple[str, str]]) -> list[str]:
    links = []
    print("Fetching Correct search result ...")
    tasks = []
    
    async def fetch_and_append_link(tuple):
        for i in range(1, 4):
            json_results = await asyncio.to_thread(make_search_request, tuple, i)
            link = compare_and_get_link(json_results, tuple)
            if link is not None:
                links.append((tuple,f"https://www.redfin.com{link}"))
                break
            if i == 3:
                print(f"Error: {tuple}")
        progress_bar(len(list_of_tuples))
    
    for tuple in list_of_tuples:
        tasks.append(fetch_and_append_link(tuple))

    await asyncio.gather(*tasks)
    return links

def get_data(link:str,query:str):
    url=f"{link}/filter/{query}"
    response = requests.get(url,impersonate='chrome110')
    html=HTMLParser(response.text)
    return html.css_first('#sidepane-header > div > div > div.descriptionAndModeContainer.flex.justify-between.align-center > div.description.flex.align-center > div > div.homes.summary').text().split(" ")[0].replace(",","")

async def scrape_data(search_links: list[str]):
    all_data=[]
    print("Scraping Data ...")
    async def get_all_data(link:str):
        result=[await asyncio.to_thread(get_data,link[1],query)for query in ["property-type=house,include=sold-1mo","property-type=house","property-type=land,include=sold-1mo","property-type=land"]]
        all_data.append((link[0],result))
        progress_bar(len(search_links))
    tasks=[get_all_data(link) for link in search_links]
    await asyncio.gather(*tasks)       
    return all_data

@timing_decorator
async def main(uploaded_file)->DataFrame:
    list_of_counties: list[str] = read_counties_from_file(uploaded_file)
    list_of_tuples: list[tuple[str, str]] = split_county_to_tuple(list_of_counties)
    search_links=await get_search_links_associated_with_search_tuple(list_of_tuples)
    progress_bar(reset=True)
    data=await scrape_data(search_links)
    data=data_to_dataframe(data)
    return data
    

if __name__ == '__main__':
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(main())
